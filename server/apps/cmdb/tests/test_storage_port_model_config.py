"""存储以太口 / FC 口模型种子契约（仅 model_config.xlsx，不含采集与布线）。

锁定：
- `storage_eth_port` / `storage_fc_port` 独立模型，归属 hardware_components，不复用 nic；
- 以太口至少有 mac、name/location，可选 ip_addr、运行状态；
- FC 口至少有 wwpn、name/location，可选 speed、运行状态；
- inst_name 建议：以太口用与 nic 相同的归一化 MAC（aa:bb:cc:dd:ee:ff），FC 口用归一化 WWPN；
- 关联：storage contains 两口（1:n）、interface connect storage_eth_port（n:n，仅预留）；
- 既有 nic 关联（interface connect nic、physcial_server contains nic）不变；
- 不引入 interface_belong_storage、rack_contains_storage，也不把 FC 口接入 interface connect。
"""

import os

import openpyxl
import pandas as pd
import pytest

pytestmark = pytest.mark.unit

XLSX = os.path.join(os.path.dirname(__file__), "..", "support-files", "model_config.xlsx")

STR_OPTION = '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}'
IPV4_OPTION = '{"validation_type": "ipv4", "custom_regex": "", "widget_type": "single_line"}'
OPERA_STATUS_OPTION = '{"enum_rule_type": "public_library", "public_library_id": "opera_status", "enum_select_mode": "single"}'

COMMON_COMPONENT_FIELDS = {
    "inst_name": "str",
    "organization": "organization",
    "self_device": "str",
    "tag": "tag",
    "name": "str",
    "location": "str",
    "running_status": "enum",
    "auto_collect": "bool",
    "collect_time": "time",
    "collect_task": "str",
}

ETH_REQUIRED_FIELDS = {
    **COMMON_COMPONENT_FIELDS,
    "mac": "str",
    "ip_addr": "str",
}

FC_REQUIRED_FIELDS = {
    **COMMON_COMPONENT_FIELDS,
    "wwpn": "str",
    "speed": "str",
}

EXISTING_NIC_ASSOCIATIONS = {
    ("interface", "nic", "connect", "n:n"),
    ("physcial_server", "nic", "contains", "1:n"),
}

NEW_PORT_ASSOCIATIONS = {
    ("storage", "storage_eth_port", "contains", "1:n"),
    ("storage", "storage_fc_port", "contains", "1:n"),
    ("interface", "storage_eth_port", "connect", "n:n"),
}

FORBIDDEN_ASSOCIATIONS = {
    ("interface", "storage", "belong", "n:1"),
    ("interface", "storage_eth_port", "belong", "n:1"),
    ("interface", "storage_fc_port", "belong", "n:1"),
    ("interface", "storage_fc_port", "connect", "n:n"),
    ("rack", "storage", "contains", "1:n"),
    ("storage_eth_port", "storage", "belong", "n:1"),
    ("storage_fc_port", "storage", "belong", "n:1"),
}


def _records(sheet):
    headers = [cell.value for cell in sheet[2]]
    return [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=3, values_only=True)
        if any(value is not None and str(value).strip() != "" for value in values)
    ]


def _attr_map(workbook, model_id):
    return {row["attr_id"]: row for row in _records(workbook[f"attr-{model_id}"])}


def _association_tuples(workbook):
    tuples = set()
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("asso-"):
            continue
        for row in _records(workbook[sheet_name]):
            tuples.add(
                (
                    row["src_model_id"],
                    row["dst_model_id"],
                    row["asst_id"],
                    row["mapping"],
                )
            )
    return tuples


def test_storage_port_models_exist_under_hardware_components():
    workbook = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    models = {row["model_id"]: row for row in _records(workbook["models"])}

    assert "storage_eth_port" in models
    assert "storage_fc_port" in models
    assert models["storage_eth_port"]["classification_id"] == "hardware_components"
    assert models["storage_fc_port"]["classification_id"] == "hardware_components"
    assert models["storage_eth_port"]["app_topo_layer"] == "infrastructure"
    assert models["storage_fc_port"]["app_topo_layer"] == "infrastructure"
    assert models["storage_eth_port"]["model_name"] == "存储以太口"
    assert models["storage_fc_port"]["model_name"] == "存储FC口"
    assert models["nic"]["model_id"] == "nic"
    assert models["storage_eth_port"]["model_id"] != "nic"
    assert models["storage_fc_port"]["model_id"] != "nic"


def test_storage_port_fields_follow_storage_and_nic_conventions():
    workbook = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    eth = _attr_map(workbook, "storage_eth_port")
    fc = _attr_map(workbook, "storage_fc_port")

    assert {key: eth[key]["attr_type"] for key in ETH_REQUIRED_FIELDS} == ETH_REQUIRED_FIELDS
    assert {key: fc[key]["attr_type"] for key in FC_REQUIRED_FIELDS} == FC_REQUIRED_FIELDS

    assert "nic_mac" not in eth
    assert "nic_mac" not in fc
    assert "wwn" not in fc
    assert eth["mac"]["attr_name"] == "MAC 地址"
    assert fc["wwpn"]["attr_name"] == "WWPN"
    assert eth["ip_addr"]["attr_name"] == "IP地址"
    assert eth["name"]["attr_name"] == "接口名称"
    assert eth["location"]["attr_name"] == "位置"
    assert fc["speed"]["attr_name"] == "速率"

    for attrs in (eth, fc):
        assert attrs["inst_name"]["is_only"] is True
        assert attrs["inst_name"]["is_required"] is True
        assert attrs["organization"]["is_required"] is True
        assert attrs["self_device"]["key_attribute"] is True
        assert attrs["running_status"]["option"] == OPERA_STATUS_OPTION
        assert attrs["name"]["option"] == STR_OPTION
        assert attrs["location"]["option"] == STR_OPTION

    assert eth["mac"]["option"] == STR_OPTION
    assert eth["ip_addr"]["option"] == IPV4_OPTION
    assert fc["wwpn"]["option"] == STR_OPTION
    assert fc["speed"]["option"] == STR_OPTION
    assert "aa:bb:cc:dd:ee:ff" in str(eth["inst_name"].get("user_prompt") or "")
    assert "WWPN" in str(fc["inst_name"].get("user_prompt") or "")


def test_model_init_xlsx_loader_sees_storage_port_sheets():
    """与 ModelMigrate.get_model_config 相同的读表方式：header=1 的全 sheet 字典。"""
    sheets = pd.read_excel(XLSX, sheet_name=None, header=1)
    model_ids = set(sheets["models"]["model_id"].dropna().astype(str))
    assert {"storage_eth_port", "storage_fc_port"} <= model_ids
    assert "attr-storage_eth_port" in sheets
    assert "attr-storage_fc_port" in sheets
    assert "asso-storage" in sheets

    def _pairs(frame):
        rows = frame[["src_model_id", "dst_model_id", "asst_id", "mapping"]].dropna(how="all")
        return {tuple(row) for row in rows.itertuples(index=False, name=None)}

    asso_storage = _pairs(sheets["asso-storage"])
    asso_interface = _pairs(sheets["asso-interface"])
    asso_server = _pairs(sheets["asso-physcial_server"])
    assert ("storage", "storage_eth_port", "contains", "1:n") in asso_storage
    assert ("storage", "storage_fc_port", "contains", "1:n") in asso_storage
    assert ("interface", "storage_eth_port", "connect", "n:n") in asso_interface
    assert ("interface", "nic", "connect", "n:n") in asso_interface
    assert ("physcial_server", "nic", "contains", "1:n") in asso_server


def test_storage_port_associations_are_seeded_and_nic_associations_unchanged():
    workbook = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    associations = _association_tuples(workbook)

    assert NEW_PORT_ASSOCIATIONS <= associations
    assert EXISTING_NIC_ASSOCIATIONS <= associations
    assert associations.isdisjoint(FORBIDDEN_ASSOCIATIONS)

    interface_connects = {(src, dst, asst, mapping) for src, dst, asst, mapping in associations if src == "interface" and asst == "connect"}
    assert ("interface", "nic", "connect", "n:n") in interface_connects
    assert ("interface", "storage_eth_port", "connect", "n:n") in interface_connects
    assert ("interface", "storage_fc_port", "connect", "n:n") not in interface_connects
